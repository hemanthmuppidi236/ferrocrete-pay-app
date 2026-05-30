"""
Excel import.

Reads an existing Pay App Excel file (G702/G703 format) and creates the
corresponding database rows: project, SOV lines, change orders, and optionally
the pay app itself.

The parser is structure-tolerant: it detects where the SOV section starts and
ends rather than hardcoding row ranges, so it works with the variety of pay-app
templates Ferrocrete projects use in the field.

Used in two modes:
  1. New project setup — user uploads the latest pay app .xlsx for an active
     project, we create the project + SOV + COs.
  2. Pay app archival — same upload, but we also create the pay app row with
     billings populated from the file's previous/this-period/stored columns.

  POST /import/pay-app-excel?create_pay_app=true
       Multipart file upload. Returns the created project (and pay app, if any).
"""

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from io import BytesIO
from openpyxl import load_workbook
from typing import Optional

from ..core.auth import CurrentUser, require_role
from ..core.supabase_client import get_service_client
from ..core import audit
from ..core.pay_app_math import save_pay_app_totals

router = APIRouter(prefix="/import", tags=["import"])


# Sheet names — these are stable across our templates and Ferrocrete's variants.
SHEET_702 = "702"
SHEET_G703 = "G703"

# Header-row markers — the SOV table's header row contains some/all of these
# strings (case-insensitive). We use these to find where data actually starts.
HEADER_MARKERS_SCHEDULED = ("scheduled value", "sched value", "scheduled")
HEADER_MARKERS_PREVIOUS = ("previous", "prev work", "from previous")
HEADER_MARKERS_DESC = ("description", "description of work")

# Phrases that indicate a "totals" or "subtotal" row — never a real SOV item.
TOTALS_PHRASES = (
    "grand total", "grand totals", "totals", "subtotal", "sub-total", "sub total",
    "total this period", "less retainage", "total earned",
)

# Phrases that indicate a section header / break — start of a new section, not a line item.
SECTION_HEADER_PHRASES = (
    "change order", "change orders", "co's", "approved change",
    "contract scope", "base contract", "scope of work",
)

# Header-info labels in column I (right side of header block).
HEADER_INFO_LABELS = {
    "project_no":      ("project no", "project #", "project number", "job no", "job #"),
    "app_no":          ("application", "app no", "app #", "invoice no", "invoice #"),
    "period_to":       ("period to", "period ending", "invoice through", "thru date"),
    "project_name":    ("project name", "project:", "project"),
}

# Sheet-702 retention rate location (Demo 1 template convention).
S702_RETENTION_CELL = "C27"


def _safe_float(v) -> float:
    if v is None or v == "" or v == " ":
        return 0.0
    if isinstance(v, str):
        if v.startswith("="):
            return 0.0    # formula not evaluated; treat as missing
        try:
            return float(v.replace(",", "").replace("$", "").strip())
        except ValueError:
            return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _safe_str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _contains_any(text: Optional[str], needles: tuple) -> bool:
    """Case-insensitive 'contains any of' check."""
    if not text:
        return False
    t = text.lower()
    return any(n in t for n in needles)


def _extract_gc_info(ws, max_search: int = 20) -> dict:
    """
    Find GC info on the 702 sheet. The convention is a 'GENERAL CONTRACTOR'
    label followed (usually directly below, in the same column) by the GC's
    company name, address lines, and optionally an email.

    Returns dict with keys: gc_company, gc_address.
    """
    info: dict = {}
    for r in range(1, max_search + 1):
        for c in range(1, 10):
            label = _safe_str(ws.cell(row=r, column=c).value)
            if not label:
                continue
            if "general contractor" in label.lower():
                # Walk down rows in the same column until we hit a blank or another label.
                lines = []
                for dr in range(1, 6):
                    v = _safe_str(ws.cell(row=r + dr, column=c).value)
                    if not v:
                        break
                    # Stop if we hit another label that looks like a heading
                    if v.endswith(":") or v.isupper() and len(v) > 5:
                        break
                    lines.append(v)
                if lines:
                    info["gc_company"] = lines[0]
                    if len(lines) > 1:
                        info["gc_address"] = "\n".join(lines[1:])
                return info
    return info


def _detect_header_row(ws, max_search: int = 30) -> Optional[int]:
    """
    Find the SOV table header row.

    Returns the 1-indexed row number whose contents include something like
    'Description' AND 'Scheduled Value', or None if no header found.
    """
    for r in range(1, max_search + 1):
        # Collect text from cols A-G
        row_text = []
        for c in range(1, 8):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                row_text.append(v.lower())
        joined = " | ".join(row_text)
        has_sched = any(m in joined for m in HEADER_MARKERS_SCHEDULED)
        has_desc = any(m in joined for m in HEADER_MARKERS_DESC)
        if has_sched and has_desc:
            return r
    return None


def _is_totals_row(desc: Optional[str]) -> bool:
    return _contains_any(desc, TOTALS_PHRASES)


def _is_section_header_row(desc: Optional[str], sched: float, prev: float, this_p: float, stored: float) -> bool:
    """
    A 'section header' has descriptive text in column B but no numeric data.
    e.g., 'Contract Scope', 'Change Orders'.
    """
    if sched != 0 or prev != 0 or this_p != 0 or stored != 0:
        return False
    return _contains_any(desc, SECTION_HEADER_PHRASES)


def _detect_co_section_start(ws, sov_end_row: int, max_search: int = 80) -> Optional[int]:
    """
    After SOV ends, look for a 'Change Order' section header row.
    The header text may appear in column A OR column B.
    Return the first row of the CO body (not the header row itself), or None.
    """
    for r in range(sov_end_row, min(sov_end_row + max_search, ws.max_row + 1)):
        a = _safe_str(ws.cell(row=r, column=1).value)
        b = _safe_str(ws.cell(row=r, column=2).value)
        for cell_text in (a, b):
            if cell_text and "change" in cell_text.lower() and (
                "order" in cell_text.lower() or "co's" in cell_text.lower()
            ):
                return r + 1   # first row after the section header
    return None


def _extract_header_info(ws, max_search: int = 12) -> dict:
    """
    Find header metadata (project no, app no, period to, project name).

    Scans rows 1..max_search for labels in cells A-I, returning the adjacent
    cell value. Robust to layout variation: some templates put the label in
    col H and value in col I; others put label and value in the same cell
    separated by a colon.
    """
    found: dict = {}
    for r in range(1, max_search + 1):
        for c in range(1, 10):
            cell_text = _safe_str(ws.cell(row=r, column=c).value)
            if not cell_text:
                continue
            text_lower = cell_text.lower().rstrip(":").strip()
            for key, labels in HEADER_INFO_LABELS.items():
                if key in found:
                    continue
                # Match: cell content is exactly the label, OR ends with the label and colon
                if any(text_lower == l or text_lower.endswith(": " + l) or text_lower.startswith(l)
                       for l in labels):
                    # Value is likely in the next cell to the right, or the one after
                    for offset in (1, 2):
                        adj = ws.cell(row=r, column=c + offset).value
                        if adj is not None and _safe_str(adj):
                            found[key] = adj
                            break
                    break
    return found


def _extract_sov_lines(ws, header_row: int) -> tuple[list, int]:
    """
    Extract SOV line items starting from the row after header_row.

    Stops at:
      - A totals/grand-total row
      - A 'Change Order' section header (returns the SOV-end row so caller knows)
      - End of sheet
      - Many consecutive blank rows (defensive)

    Returns (sov_list, last_row_consumed).
    """
    sov_lines = []
    blank_streak = 0
    last_row = header_row
    in_change_order_section = False

    for r in range(header_row + 1, ws.max_row + 1):
        item_no = _safe_str(ws.cell(row=r, column=1).value)
        desc = _safe_str(ws.cell(row=r, column=2).value)
        sched = _safe_float(ws.cell(row=r, column=3).value)
        prev = _safe_float(ws.cell(row=r, column=4).value)
        this_p = _safe_float(ws.cell(row=r, column=5).value)
        stored = _safe_float(ws.cell(row=r, column=6).value)

        # Skip totally-empty rows but count them
        if not desc and not item_no and sched == 0 and prev == 0 and this_p == 0 and stored == 0:
            blank_streak += 1
            if blank_streak >= 8:
                break    # likely past the data
            continue
        blank_streak = 0
        last_row = r

        # Detect "Change Orders" section header in column A (some templates put
        # it there instead of column B). End SOV when we hit it.
        if item_no and "change" in item_no.lower() and (
            "order" in item_no.lower() or "co's" in item_no.lower()
        ):
            break

        # Skip totals rows entirely (don't import "GRAND TOTALS" as an SOV line)
        if _is_totals_row(desc):
            # GRAND TOTALS typically marks the end of the document body.
            break

        # Section headers — recognize the boundary, don't import as line items
        if _is_section_header_row(desc, sched, prev, this_p, stored):
            if desc and "change" in desc.lower():
                in_change_order_section = True
                break    # SOV is done; caller will pick up CO section
            else:
                # e.g., "Contract Scope" header — just a label, skip
                continue

        # If we're in the change-order section, stop reading SOV
        if in_change_order_section:
            break

        # Real SOV line: must have a description AND either a scheduled value or prior billing.
        # (Skip rows that have only stray data and no description.)
        if not desc:
            continue

        sov_lines.append({
            "item_no": item_no,
            "description": desc,
            "scheduled_value": sched,
            "billing": {
                "previous_work": prev,
                "this_period_work": this_p,
                "materials_stored": stored,
            },
            "_row": r,
        })

    return sov_lines, last_row


def _extract_co_lines(ws, start_row: int) -> list:
    """
    Extract change order lines starting from the row after the 'Change Orders'
    section header. Stops at the first totals row or many consecutive blanks.
    """
    cos = []
    blank_streak = 0
    co_index = 0

    for r in range(start_row, ws.max_row + 1):
        co_no = _safe_str(ws.cell(row=r, column=1).value)
        desc = _safe_str(ws.cell(row=r, column=2).value)
        amount = _safe_float(ws.cell(row=r, column=3).value)
        prev = _safe_float(ws.cell(row=r, column=4).value)
        this_p = _safe_float(ws.cell(row=r, column=5).value)
        stored = _safe_float(ws.cell(row=r, column=6).value)

        if not desc and not co_no and amount == 0 and prev == 0 and this_p == 0 and stored == 0:
            blank_streak += 1
            if blank_streak >= 8:
                break
            continue
        blank_streak = 0

        if _is_totals_row(desc):
            break

        if not desc:
            continue

        co_index += 1
        cos.append({
            "co_no": co_no or f"CO-{co_index:03d}",
            "description": desc,
            "amount": amount,
            "billing": {
                "previous_work": prev,
                "this_period_work": this_p,
                "materials_stored": stored,
            },
            "_row": r,
        })

    return cos


@router.post("/pay-app-excel")
async def import_pay_app_excel(
    file: UploadFile = File(...),
    create_pay_app: bool = Query(False, description="Also create the pay app row from the file's billings"),
    user: CurrentUser = Depends(require_role("admin", "accountant", "pe")),
):
    """Read a pay app .xlsx and create project + SOV + COs in the database."""
    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Failed to read Excel file: {e}")

    if SHEET_G703 not in wb.sheetnames:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"File must contain a '{SHEET_G703}' sheet (the continuation sheet with the SOV table)",
        )

    g703 = wb[SHEET_G703]
    s702 = wb[SHEET_702] if SHEET_702 in wb.sheetnames else None

    # ─── Find header row ────────────────────────────────────────────────
    header_row = _detect_header_row(g703)
    if not header_row:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Could not find SOV table header row. Expected a row containing "
            "'Description' and 'Scheduled Value' within the first 30 rows of G703.",
        )

    # ─── Extract project metadata ────────────────────────────────────────
    # Try the G703 header first, then the 702 sheet (different templates put
    # the info in different places).
    header_info = _extract_header_info(g703)
    if s702 is not None:
        # 702 sheet has more structured header info; merge in any missing fields
        s702_info = _extract_header_info(s702)
        for k, v in s702_info.items():
            header_info.setdefault(k, v)

    # GC info also lives in sheet 702 in many templates: row with the company
    # name follows the "GENERAL CONTRACTOR" label.
    gc_info = _extract_gc_info(s702) if s702 is not None else {}

    project_no = _safe_str(header_info.get("project_no"))
    app_no = header_info.get("app_no")
    period_to = header_info.get("period_to")
    project_name = _safe_str(header_info.get("project_name"))

    # Retention rate from 702 sheet if available
    retention_rate = 0.10
    if s702 is not None:
        rv = _safe_float(s702[S702_RETENTION_CELL].value)
        if rv:
            retention_rate = rv

    if not project_no:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Could not extract project number. Expected a 'Project No' label in the header area of G703.",
        )

    if not project_name and file.filename:
        # Try to derive from filename: "26-01_-_The_Cleo_-_Draw_2604_R2.xlsx"
        # Strip extension and project number prefix
        stem = file.filename.rsplit(".", 1)[0]
        # Remove leading "PROJECT_NO_-_"
        if project_no in stem:
            stem = stem.split(project_no, 1)[-1].lstrip("_- ")
        # Remove draw/period suffix
        for sep in ["_-_Draw", "_-_DRAW", "_-_draw"]:
            if sep in stem:
                stem = stem.split(sep)[0]
                break
        project_name = stem.replace("_", " ").strip() or f"Project {project_no}"
    if not project_name:
        project_name = f"Project {project_no}"

    # ─── Extract SOV lines ───────────────────────────────────────────────
    sov_lines, sov_end_row = _extract_sov_lines(g703, header_row)

    # ─── Extract CO lines (look for section header after SOV) ───────────
    co_start = _detect_co_section_start(g703, sov_end_row)
    cos = _extract_co_lines(g703, co_start) if co_start else []

    if not sov_lines:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Could not find any SOV line items in the G703 sheet. "
            "Make sure the file has a 'Description' / 'Scheduled Value' header row "
            "followed by line items.",
        )

    # Compute contract value as sum of SOV scheduled_value
    contract_value = sum(s["scheduled_value"] for s in sov_lines)

    sb = get_service_client()

    # Check for existing project (including soft-deleted ones, so we can revive them
    # rather than failing on the project_no UNIQUE constraint).
    existing = sb.table("projects").select("*").eq("project_no", project_no).limit(1).execute()

    project = None
    action = None

    if existing.data:
        existing_row = existing.data[0]
        if existing_row.get("deleted_at"):
            # Project was soft-deleted — revive it and treat as a fresh import.
            # Clear stale child data first (pay apps, billings, SOV lines, COs)
            # since the import will recreate them all from the file.
            revive_id = existing_row["id"]
            # Order matters: pay_app_billings has FKs to pay_apps and to
            # sov_lines/change_orders. Delete it first (via each pay_app's id),
            # then the parents.
            pa_rows = sb.table("pay_apps").select("id").eq("project_id", revive_id).execute()
            for pa_row in (pa_rows.data or []):
                sb.table("pay_app_billings").delete().eq("pay_app_id", pa_row["id"]).execute()
            sb.table("pay_apps").delete().eq("project_id", revive_id).execute()
            sb.table("change_orders").delete().eq("project_id", revive_id).execute()
            sb.table("sov_lines").delete().eq("project_id", revive_id).execute()

            # Now update the project row: clear deleted_at, refresh fields from file.
            update_payload = {
                "deleted_at": None,
                "name": project_name,
                "contract_value": str(contract_value),
                "retention_rate": str(retention_rate),
                "status": "active",
            }
            if gc_info.get("gc_company"):
                update_payload["gc_company"] = gc_info["gc_company"]
            if gc_info.get("gc_address"):
                update_payload["gc_address"] = gc_info["gc_address"]
            res = sb.table("projects").update(update_payload).eq("id", revive_id).execute()
            project = res.data[0] if res.data else {**existing_row, **update_payload}
            action = "revived"
            audit.log(user.id, "project", project["id"], "revived_via_import",
                      before=existing_row, after=project,
                      metadata={"source_file": file.filename})
        else:
            # Active existing project — match it, don't duplicate.
            project = existing_row
            action = "matched_existing"
    else:
        proj_payload = {
            "project_no": project_no,
            "name": project_name,
            "contract_value": str(contract_value),
            "retention_rate": str(retention_rate),
            "status": "active",
            "created_by": user.id,
        }
        if gc_info.get("gc_company"):
            proj_payload["gc_company"] = gc_info["gc_company"]
        if gc_info.get("gc_address"):
            proj_payload["gc_address"] = gc_info["gc_address"]
        res = sb.table("projects").insert(proj_payload).execute()
        project = res.data[0]
        action = "created"
        audit.log(user.id, "project", project["id"], "imported_from_excel",
                  after=project, metadata={"source_file": file.filename})

    # Insert SOV lines + COs for new OR revived projects (both need fresh child data).
    if action in ("created", "revived"):
        sov_payload = [{
            "project_id": project["id"],
            "item_no": s["item_no"],
            "description": s["description"],
            "scheduled_value": str(s["scheduled_value"]),
            "sort_order": i,
        } for i, s in enumerate(sov_lines)]
        if sov_payload:
            sov_res = sb.table("sov_lines").insert(sov_payload).execute()
            for i, s in enumerate(sov_lines):
                s["_db_id"] = sov_res.data[i]["id"]

        co_payload = [{
            "project_id": project["id"],
            "co_no": c["co_no"],
            "description": c["description"],
            "amount": str(c["amount"]),
            "status": "approved",
            "has_retention": True,
        } for c in cos]
        if co_payload:
            co_res = sb.table("change_orders").insert(co_payload).execute()
            for i, c in enumerate(cos):
                c["_db_id"] = co_res.data[i]["id"]
    else:
        # For existing projects, look up SOV/CO ids by description match (best-effort)
        sov_existing = sb.table("sov_lines").select("*").eq("project_id", project["id"]).execute()
        sov_by_desc = {s["description"]: s["id"] for s in sov_existing.data}
        for s in sov_lines:
            s["_db_id"] = sov_by_desc.get(s["description"])
        co_existing = sb.table("change_orders").select("*").eq("project_id", project["id"]).execute()
        co_by_no = {c["co_no"]: c["id"] for c in co_existing.data}
        for c in cos:
            c["_db_id"] = co_by_no.get(c["co_no"])

    result = {
        "project": project,
        "action": action,
        "sov_count": len(sov_lines),
        "co_count": len(cos),
    }

    # Optionally create the pay app row
    if create_pay_app and app_no:
        from datetime import date, datetime
        import calendar
        import re

        # Step 1: Derive `period` (YY-MM) — prefer filename, fall back to period_to.
        period = None
        if file.filename:
            matches = re.findall(r"(\d{2})-(\d{2})", file.filename)
            for yy, mm in reversed(matches):
                if 1 <= int(mm) <= 12:
                    period = f"{yy}-{mm}"
                    break

        # Step 2: Normalize period_to to a `date` object. We need this regardless
        # of where `period` came from, since the DB column is NOT NULL.
        pdate = None
        if period_to is not None:
            try:
                if isinstance(period_to, datetime):
                    pdate = period_to.date()
                elif isinstance(period_to, date):
                    pdate = period_to
                elif isinstance(period_to, str):
                    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d",
                                "%m-%d-%Y", "%B %d, %Y", "%b %d, %Y"):
                        try:
                            pdate = datetime.strptime(period_to.strip(), fmt).date()
                            break
                        except ValueError:
                            continue
            except Exception as e:
                print(f"[import] period_to parse failed (non-fatal): {e}", flush=True)

        # If we got a date from period_to and didn't get period from filename,
        # derive period from the date.
        if not period and pdate:
            period = f"{pdate.year % 100:02d}-{pdate.month:02d}"

        # If we got period from filename but no usable period_to, derive period_to
        # as the last day of that month (the AIA G702 convention).
        if period and not pdate:
            try:
                yy_s, mm_s = period.split("-")
                yr = 2000 + int(yy_s)
                mo = int(mm_s)
                last_day = calendar.monthrange(yr, mo)[1]
                pdate = date(yr, mo, last_day)
                print(
                    f"[import] period_to derived from period: {pdate.isoformat()} "
                    f"(filename had period {period} but no usable PERIOD TO date in the file)",
                    flush=True,
                )
            except Exception as e:
                print(f"[import] failed to derive period_to from period: {e}", flush=True)

        if not period:
            return {
                **result,
                "warning": (
                    "Pay app NOT created: could not determine period from the file. "
                    "Expected a date in 'Period To' / 'Invoice Through' (e.g. '4/30/2026') "
                    "or a YY-MM pattern in the filename. The project, SOV, and COs were "
                    "imported successfully — you'll need to create the pay app manually."
                ),
            }
        if not pdate:
            return {
                **result,
                "warning": (
                    "Pay app NOT created: could not determine the period-ending date. "
                    "Expected a date in 'PERIOD TO' / 'INVOICE THROUGH' on the G703 or 702 "
                    "sheet (e.g. '4/30/2026'). The project, SOV, and COs were imported "
                    "successfully — you'll need to create the pay app manually."
                ),
            }

        # Check for existing
        pa_existing = (sb.table("pay_apps")
                       .select("id").eq("project_id", project["id"])
                       .eq("period", period).limit(1).execute())
        if pa_existing.data:
            return {**result, "pay_app_id": pa_existing.data[0]["id"], "pay_app_action": "exists"}

        pa_payload = {
            "project_id": project["id"],
            "period": period,
            "app_no": int(app_no) if isinstance(app_no, (int, float)) else 1,
            "period_to": pdate.isoformat(),
            "status": "draft",
        }
        pa_res = sb.table("pay_apps").insert(pa_payload).execute()
        pa = pa_res.data[0]

        # Insert billings
        billing_rows = []
        for s in sov_lines:
            if s.get("_db_id"):
                billing_rows.append({
                    "pay_app_id": pa["id"],
                    "sov_line_id": s["_db_id"],
                    "previous_work": str(s["billing"]["previous_work"]),
                    "this_period_work": str(s["billing"]["this_period_work"]),
                    "materials_stored": str(s["billing"]["materials_stored"]),
                })
        for c in cos:
            if c.get("_db_id"):
                billing_rows.append({
                    "pay_app_id": pa["id"],
                    "change_order_id": c["_db_id"],
                    "previous_work": str(c["billing"]["previous_work"]),
                    "this_period_work": str(c["billing"]["this_period_work"]),
                    "materials_stored": str(c["billing"]["materials_stored"]),
                })
        if billing_rows:
            sb.table("pay_app_billings").insert(billing_rows).execute()

        # Recalc totals (best-effort; don't fail import if this errors).
        # previous_certificates is derived from each line's previous_work
        # column (the per-line carry-forward), so mid-project imports compute
        # correctly without needing a 702!G29 override.
        try:
            save_pay_app_totals(pa["id"])
        except Exception as e:
            print(f"[import] save_pay_app_totals failed: {e}", flush=True)

        # Re-read the pay app row so we have the final totals for the
        # release tracker auto-create.
        try:
            fresh_pa = (sb.table("pay_apps").select("*")
                        .eq("id", pa["id"]).limit(1).execute())
            if fresh_pa.data:
                from .pay_apps import _autocreate_release_tracker
                _autocreate_release_tracker(sb, fresh_pa.data[0])
        except Exception as e:
            print(f"[import] release tracker auto-create failed (non-fatal): {e}", flush=True)

        audit.log(user.id, "pay_app", pa["id"], "imported_from_excel",
                  after=pa, metadata={"source_file": file.filename})
        result["pay_app_id"] = pa["id"]
        result["pay_app_action"] = "created"

    return result

