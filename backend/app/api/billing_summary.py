"""
Billing Period Summary API (internal monthly roll-up).

  GET   /billing-summary?period=YY-MM   one row per project for the period,
                                        with period totals + accrued-to-date
  PATCH /billing-summary/override       upsert the manual columns for a cell

Auto columns come from pay apps (contract/completed/retention/billed) and
release trackers (Potential Net = Ferrocrete's net income, and CP/CF·UP/UF
sent flags). Only the free-text/manual columns are stored, in
billing_summary_overrides (migration 004).
"""

from decimal import Decimal
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from typing import Optional

from ..core.auth import CurrentUser, get_current_user, require_role
from ..core.supabase_client import get_service_client
from ..core import audit
from ..core import billing_math as bm
from ..schemas.billing import (
    BillingSummaryResponse, BillingOverrideUpdate, QuickbooksUpdate,
)

router = APIRouter(prefix="/billing-summary", tags=["billing_summary"])


def _sum_by(rows, key_field, val_field) -> dict:
    """Sum val_field grouped by key_field, as Decimals."""
    out: dict = {}
    for r in rows or []:
        k = r.get(key_field)
        out[k] = out.get(k, Decimal("0")) + bm.dec(r.get(val_field))
    return out


def _load_overrides(sb, period: str) -> dict:
    """Overrides keyed by project_id. Degrades to {} if migration 004 hasn't
    been applied yet, so the summary still works from auto sources alone."""
    try:
        res = (sb.table("billing_summary_overrides")
               .select("*").eq("period", period).execute())
        return {r["project_id"]: r for r in (res.data or [])}
    except Exception as e:
        print(f"[billing-summary] overrides unavailable (run migration 004?): {e}", flush=True)
        return {}


@router.get("", response_model=BillingSummaryResponse)
def get_billing_summary(
    period: Optional[str] = Query(None, description="YY-MM; defaults to latest"),
    user: CurrentUser = Depends(get_current_user),
):
    sb = get_service_client()

    # ─── Available periods (union of pay apps + trackers) ───
    pa_periods = {r["period"] for r in (sb.table("pay_apps").select("period").execute().data or []) if r.get("period")}
    rt_periods = {r["period"] for r in (sb.table("release_trackers").select("period").execute().data or []) if r.get("period")}
    available = sorted(pa_periods | rt_periods, reverse=True)

    if not period:
        period = available[0] if available else None
    if not period:
        # Empty system — return an empty but well-formed payload.
        zero = Decimal("0")
        return {
            "period": None, "available_periods": [], "rows": [],
            "totals": {k: zero for k in
                       ["revised_contract", "total_completed", "retention",
                        "balance_to_finish", "gross_billing", "billed_amount",
                        "potential_net", "rebar", "cmu"]},
            "accrued": {"net": zero, "billed": zero},
            "footer": {"net_pct_of_billed": None, "quickbooks_total": None,
                       "quickbooks_diff": None, "running_total_billed": zero},
        }

    # ─── Projects (non-deleted) ───
    try:
        proj_rows = sb.table("projects").select(
            "id, project_no, name, contract_value, retention_rate, gc_contact_email, "
            "billing_due_rule, billing_contact, status, deleted_at"
        ).execute().data or []
    except Exception:
        # Migration 007 not applied yet — fall back without the new columns.
        proj_rows = sb.table("projects").select(
            "id, project_no, name, contract_value, retention_rate, "
            "gc_contact_email, status, deleted_at"
        ).execute().data or []
    projects = {p["id"]: p for p in proj_rows if not p.get("deleted_at")}

    # Approved change-order totals (revised-contract fallback when no pay app)
    co_by_project = _sum_by(
        sb.table("change_orders").select("project_id, amount, status")
          .eq("status", "approved").execute().data,
        "project_id", "amount",
    )

    # ─── Pay apps up to and including the period ───
    pay_apps_le = sb.table("pay_apps").select(
        "id, project_id, period, due_date, status, revised_contract, "
        "total_completed_to_date, retention_held, current_payment_due"
    ).lte("period", period).execute().data or []
    payapp_by_project = {r["project_id"]: r for r in pay_apps_le if r.get("period") == period}
    accrued_billed = sum((bm.dec(r.get("current_payment_due")) for r in pay_apps_le), Decimal("0"))

    # ─── Release trackers up to and including the period ───
    trackers_le = sb.table("release_trackers").select(
        "id, project_id, period, invoice_amount"
    ).lte("period", period).execute().data or []
    tracker_ids = [t["id"] for t in trackers_le]

    # Check + unbilled sums per tracker (batched)
    checks_by_tracker: dict = {}
    unbilled_by_tracker: dict = {}
    if tracker_ids:
        checks_by_tracker = _sum_by(
            sb.table("release_lines").select("release_tracker_id, check_amount")
              .in_("release_tracker_id", tracker_ids).execute().data,
            "release_tracker_id", "check_amount",
        )
        unbilled_by_tracker = _sum_by(
            sb.table("release_unbilled_entries").select("release_tracker_id, amount")
              .in_("release_tracker_id", tracker_ids).execute().data,
            "release_tracker_id", "amount",
        )

    # Net per tracker; accrued net = Σ over all periods ≤ selected (skip no-invoice)
    net_by_tracker: dict = {}
    accrued_net = Decimal("0")
    period_tracker_by_project: dict = {}
    period_tracker_ids = []
    for t in trackers_le:
        n = bm.ferrocrete_net(
            t.get("invoice_amount"),
            checks_by_tracker.get(t["id"], Decimal("0")),
            unbilled_by_tracker.get(t["id"], Decimal("0")),
        )
        net_by_tracker[t["id"]] = n
        if n is not None:
            accrued_net += n
        if t.get("period") == period:
            period_tracker_by_project[t["project_id"]] = t
            period_tracker_ids.append(t["id"])

    # ─── Waiver auto-flags (P/Q) for this period's trackers ───
    waiver_flags_by_tracker: dict = {}
    if period_tracker_ids:
        lines = (sb.table("release_lines").select("id, release_tracker_id")
                 .in_("release_tracker_id", period_tracker_ids).execute().data or [])
        tracker_by_line = {ln["id"]: ln["release_tracker_id"] for ln in lines}
        if tracker_by_line:
            waivers = (sb.table("waivers").select("release_line_id, waiver_type")
                       .in_("release_line_id", list(tracker_by_line.keys())).execute().data or [])
            for w in waivers:
                tid = tracker_by_line.get(w["release_line_id"])
                if not tid:
                    continue
                flags = waiver_flags_by_tracker.setdefault(tid, {"cp_cf": False, "up_uf": False})
                if w.get("waiver_type") in ("CP", "CF"):
                    flags["cp_cf"] = True
                elif w.get("waiver_type") in ("UP", "UF"):
                    flags["up_uf"] = True

    overrides = _load_overrides(sb, period)

    # ─── Inclusion set: active projects ∪ any project with activity this period ─
    include_ids = {pid for pid, p in projects.items() if (p.get("status") or "active") == "active"}
    include_ids |= set(payapp_by_project.keys())
    include_ids |= set(period_tracker_by_project.keys())
    include_ids &= set(projects.keys())   # only known, non-deleted projects

    rows = []
    for pid in include_ids:
        tracker = period_tracker_by_project.get(pid)
        rows.append(bm.assemble_row(
            project=projects[pid],
            pay_app=payapp_by_project.get(pid),
            net=net_by_tracker.get(tracker["id"]) if tracker else None,
            override=overrides.get(pid, {}),
            co_total=co_by_project.get(pid, Decimal("0")),
            waiver_flags=waiver_flags_by_tracker.get(tracker["id"]) if tracker else {},
        ))

    # Projects flagged "Skip" in Billing Due Date sort to the bottom.
    rows.sort(key=lambda r: (bm.is_skip(r.get("billing_due_date")),
                             r["project_no"], r["project_name"]))

    totals = bm.summarize_totals(rows)

    # ─── Footer reconciliation ───
    qb_total = None
    try:
        meta = (sb.table("billing_period_meta").select("quickbooks_total")
                .eq("period", period).limit(1).execute())
        if meta.data and meta.data[0].get("quickbooks_total") is not None:
            qb_total = bm.dec(meta.data[0]["quickbooks_total"])
    except Exception:
        pass  # migration 007 not applied — no Quickbooks total yet

    billed_total = totals["billed_amount"]
    net_total = totals["potential_net"]
    net_pct = (net_total / billed_total) if billed_total else None
    qb_diff = (qb_total - billed_total) if qb_total is not None else None

    # Running total billed = Σ current_payment_due across the selected year,
    # for periods up to and including this one.
    year = period.split("-")[0]
    running_billed = sum(
        (bm.dec(r.get("current_payment_due")) for r in pay_apps_le
         if (r.get("period") or "").split("-")[0] == year),
        Decimal("0"),
    )

    return {
        "period": period,
        "available_periods": available,
        "rows": rows,
        "totals": totals,
        "accrued": {"net": accrued_net, "billed": accrued_billed},
        "footer": {
            "net_pct_of_billed": net_pct,
            "quickbooks_total": qb_total,
            "quickbooks_diff": qb_diff,
            "running_total_billed": running_billed,
        },
    }


@router.patch("/override")
def upsert_override(
    body: BillingOverrideUpdate,
    user: CurrentUser = Depends(require_role("admin", "accountant", "pe")),
):
    """Upsert the manual columns for one (project, period). Only the fields sent
    are written; unset fields are left untouched."""
    sb = get_service_client()

    proj = sb.table("projects").select("id").eq("id", str(body.project_id)).limit(1).execute()
    if not proj.data:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Project not found")

    fields = body.model_dump(mode="json", exclude_unset=True, exclude={"project_id", "period"})

    try:
        existing = (sb.table("billing_summary_overrides").select("id")
                    .eq("project_id", str(body.project_id)).eq("period", body.period)
                    .limit(1).execute())
        if existing.data:
            res = (sb.table("billing_summary_overrides").update(fields)
                   .eq("id", existing.data[0]["id"]).execute())
        else:
            payload = {"project_id": str(body.project_id), "period": body.period, **fields}
            res = sb.table("billing_summary_overrides").insert(payload).execute()
    except Exception as e:
        raise HTTPException(
            status.HTTP_503_SERVICE_UNAVAILABLE,
            f"Billing overrides unavailable — has migration 004 been applied? ({type(e).__name__}: {e})",
        )

    row = res.data[0] if res.data else {}
    audit.log(user.id, "billing_summary_override", row.get("id", ""), "upserted",
              after=row, metadata={"project_id": str(body.project_id), "period": body.period})
    return row


@router.patch("/quickbooks")
def upsert_quickbooks(
    body: QuickbooksUpdate,
    user: CurrentUser = Depends(require_role("admin", "accountant", "pe")),
):
    """Set the per-period Quickbooks total for the footer reconciliation."""
    sb = get_service_client()
    payload = {"period": body.period,
               "quickbooks_total": str(body.quickbooks_total)
               if body.quickbooks_total is not None else None}
    try:
        existing = (sb.table("billing_period_meta").select("period")
                    .eq("period", body.period).limit(1).execute())
        if existing.data:
            res = (sb.table("billing_period_meta")
                   .update({"quickbooks_total": payload["quickbooks_total"]})
                   .eq("period", body.period).execute())
        else:
            res = sb.table("billing_period_meta").insert(payload).execute()
    except Exception as e:
        raise HTTPException(
            status.HTTP_503_SERVICE_UNAVAILABLE,
            f"Quickbooks total unavailable — has migration 007 been applied? "
            f"({type(e).__name__}: {e})",
        )
    audit.log(user.id, "billing_period_meta", body.period, "quickbooks_set",
              metadata={"period": body.period})
    return res.data[0] if res.data else payload


# Columns for the Excel export, in the reference-sheet order (18 columns).
_EXPORT_COLS = [
    ("Job", "job", "text"),
    ("Billing Due Date", "billing_due_date", "text"),
    ("Revised Contract", "revised_contract", "money"),
    ("Total Complete and Stored to Date", "total_completed", "money"),
    ("Retention", "retention", "money"),
    ("Balance to Finish", "balance_to_finish", "money"),
    ("Balance to Finish W/Ret", "balance_with_retention", "money"),
    ("Gross Billing", "gross_billing", "money"),
    ("Retention %", "retention_rate", "pct"),
    ("Billed Amount", "billed_amount", "money"),
    ("Potential Net", "potential_net", "money"),
    ("BT", "bt_note", "text"),
    ("Billing Contact", "billing_contact", "text"),
    ("Payment/Billing Status", "payment_status", "text"),
]
_TOTALS_KEYS = {
    "revised_contract", "total_completed", "retention", "balance_to_finish",
    "gross_billing", "billed_amount", "potential_net", "rebar", "cmu",
}


@router.get("/export.xlsx")
def export_billing_summary(
    period: Optional[str] = Query(None),
    user: CurrentUser = Depends(get_current_user),
):
    """Excel export of the summary, matching the reference sheet's columns."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    data = get_billing_summary(period=period, user=user)
    period = data.get("period") or period or "summary"
    rows = data.get("rows") or []
    totals = data.get("totals") or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Billing Summary"

    MONEY = "$#,##0.00"
    PCT = "0%"
    bold = Font(bold=True)

    # Header
    for c, (label, _key, _kind) in enumerate(_EXPORT_COLS, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = bold

    # Data rows
    r = 2
    for row in rows:
        for c, (_label, key, kind) in enumerate(_EXPORT_COLS, start=1):
            val = row.get(key)
            cell = ws.cell(row=r, column=c)
            if kind in ("money", "pct") and val is not None:
                cell.value = float(val)
                cell.number_format = MONEY if kind == "money" else PCT
            else:
                cell.value = "" if val is None else val
        r += 1

    # Totals row (bold), summing E, F, G, H, J, L, M, O, P.
    ws.cell(row=r, column=1, value="Total").font = bold
    for c, (_label, key, kind) in enumerate(_EXPORT_COLS, start=1):
        if key in _TOTALS_KEYS and totals.get(key) is not None:
            cell = ws.cell(row=r, column=c, value=float(totals[key]))
            cell.number_format = MONEY
            cell.font = bold

    # Column widths + freeze header
    ws.column_dimensions["A"].width = 34
    for c in range(2, len(_EXPORT_COLS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Billing_Summary_{period}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
