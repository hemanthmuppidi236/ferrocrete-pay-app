"""
Release tracker Excel generation.

Uses the existing release_engine.populate_sheet() helper to do the heavy lifting
(sub-tier merge handling, downstream formula re-pointing). We just bridge from
DB rows into the dict shape it expects.
"""

from io import BytesIO
from pathlib import Path
from datetime import datetime, timezone, date
import calendar

from openpyxl import load_workbook

from .config import settings
from .supabase_client import get_service_client
from . import storage as storage_helpers


# Path to the Excel template inside the backend repo
TEMPLATE_DIR = Path(__file__).resolve().parents[2] / "engines" / "templates"
RELEASE_TEMPLATE = TEMPLATE_DIR / "ReleaseTracker_Template.xlsx"


def _build_subs_payload(tracker_id: str) -> list:
    """Build the subs list shape that release_engine.populate_sheet expects.

    Returns a list of:
      {name, release_type, exception, billed, check, prev_month_status,
       sub_tiers: [{name, release_type, exception, check, prev_month_status}, ...],
       non_prelimed: bool}
    """
    sb = get_service_client()

    # Load all release_lines for this tracker, joined with sub data
    lines_res = (sb.table("release_lines")
                 .select("*, subs(id, name, parent_sub_id, is_non_prelimed, sort_order)")
                 .eq("release_tracker_id", str(tracker_id))
                 .execute())

    # Index by sub id for quick lookup
    line_by_sub = {ln["subs"]["id"]: ln for ln in lines_res.data}

    # Group: parents first (parent_sub_id = NULL), then sub-tiers grouped under their parent
    parents = []
    children_by_parent = {}
    for ln in lines_res.data:
        sub = ln["subs"]
        if sub["parent_sub_id"] is None:
            parents.append((sub, ln))
        else:
            children_by_parent.setdefault(sub["parent_sub_id"], []).append((sub, ln))

    # Sort parents by sort_order then name
    parents.sort(key=lambda x: (x[0].get("sort_order") or 0, x[0]["name"]))

    payload = []
    for sub, ln in parents:
        sub_tiers = []
        for child_sub, child_ln in sorted(
            children_by_parent.get(sub["id"], []),
            key=lambda x: (x[0].get("sort_order") or 0, x[0]["name"]),
        ):
            sub_tiers.append({
                "name": child_sub["name"],
                "release_type": child_ln.get("release_type"),
                "exception": child_ln.get("exception"),
                "check": float(child_ln.get("check_amount") or 0),
                "billed": float(child_ln.get("billed_amount") or 0),
                "prev_month_status": child_ln.get("prev_month_status"),
            })

        payload.append({
            "name": sub["name"],
            "release_type": ln.get("release_type"),
            "exception": ln.get("exception"),
            "billed": float(ln.get("billed_amount") or 0),
            "check": float(ln.get("check_amount") or 0),
            "prev_month_status": ln.get("prev_month_status"),
            "sub_tiers": sub_tiers,
            "non_prelimed": sub.get("is_non_prelimed", False),
        })

    return payload


def generate_release_tracker_excel(tracker_id: str) -> bytes:
    """Generate the release tracker .xlsx as bytes."""
    if not RELEASE_TEMPLATE.exists():
        raise FileNotFoundError(f"Release template missing at {RELEASE_TEMPLATE}")

    sb = get_service_client()

    # Load tracker + project
    rt_res = sb.table("release_trackers").select("*, projects(*)").eq("id", str(tracker_id)).limit(1).execute()
    if not rt_res.data:
        raise ValueError(f"Release tracker {tracker_id} not found")
    rt = rt_res.data[0]
    project = rt["projects"]

    # Build the subs payload
    subs = _build_subs_payload(tracker_id)

    # Load template
    with open(RELEASE_TEMPLATE, "rb") as f:
        wb = load_workbook(BytesIO(f.read()))

    # Rename TEMPLATE sheet to the period
    period = rt["period"]
    if "TEMPLATE" in wb.sheetnames:
        ws = wb["TEMPLATE"]
        ws.title = period
    elif period in wb.sheetnames:
        ws = wb[period]
    else:
        # Fallback: take the first sheet
        ws = wb[wb.sheetnames[0]]
        ws.title = period

    # Use the existing release_engine.populate_sheet to do the heavy lifting
    # (sub-tier merges, downstream formula re-pointing, section styles)
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "engines"))
    import release_engine as RE

    project_info = {"name": project["name"], "project_no": project["project_no"]}
    invoice_amount = float(rt["invoice_amount"]) if rt.get("invoice_amount") else None

    RE.populate_sheet(ws, project_info, subs, period, invoice_amount=invoice_amount)
    RE.apply_protection(ws)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_and_store_release_tracker_excel(tracker_id: str) -> dict:
    sb = get_service_client()
    rt_res = sb.table("release_trackers").select("*, projects(project_no, name)").eq("id", str(tracker_id)).limit(1).execute()
    if not rt_res.data:
        raise ValueError(f"Release tracker {tracker_id} not found")
    rt = rt_res.data[0]
    project = rt["projects"]

    excel_bytes = generate_release_tracker_excel(tracker_id)
    path = storage_helpers.make_release_tracker_path(
        rt["period"], project["project_no"], project["name"],
    )
    storage_helpers.upload_bytes(
        settings.bucket_release_trackers,
        path,
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    now = datetime.now(timezone.utc).isoformat()
    sb.table("release_trackers").update({
        "excel_file_path": path,
        "excel_generated_at": now,
    }).eq("id", str(tracker_id)).execute()

    return {
        "file_path": path,
        "download_url": storage_helpers.signed_url(settings.bucket_release_trackers, path),
        "generated_at": now,
    }
