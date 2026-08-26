"""
WI-6: Release tracker Excel export, laid out like the reference tracker sheet
(2935 A Street, sheet 26-06): header block, Subs/Vendors table with a Difference
column, Subs total, Ferrocrete Total, Non-Prelimed Bills, Previous Month(s)
Unbilled, Ferrocrete Net, and the Buildertrend reconciliation with Discrepancy.

Totals are Excel formulas, not baked values, so the sheet stays live if edited.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .supabase_client import get_service_client
from . import release_stage as rs

MONEY = "$#,##0.00"
_BOLD = Font(bold=True)
_SHADE = PatternFill("solid", fgColor="EFE7D2")
_THIN = Side(style="thin", color="C9BC9C")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _period_to(period: str) -> str:
    try:
        import calendar
        yy, mm = period.split("-")
        last = calendar.monthrange(2000 + int(yy), int(mm))[1]
        return f"20{yy}-{mm}-{last:02d}"
    except Exception:
        return period


def generate_tracker_export_xlsx(tracker_id: str) -> tuple[bytes, str]:
    """Build the tracker export workbook; returns (bytes, filename)."""
    sb = get_service_client()
    rt = sb.table("release_trackers").select("*").eq("id", str(tracker_id)).limit(1).execute()
    if not rt.data:
        raise ValueError("Release tracker not found")
    t = rt.data[0]
    proj = sb.table("projects").select("project_no, name").eq("id", t["project_id"]).limit(1).execute()
    project = proj.data[0] if proj.data else {}

    lines_res = (sb.table("release_lines")
                 .select("*, subs(name, is_non_prelimed)")
                 .eq("release_tracker_id", str(tracker_id)).execute())
    prelim, nonprelim, lwn = [], [], []
    for ln in (lines_res.data or []):
        sub = ln.pop("subs", None) or {}
        ln["sub_name"] = sub.get("name") or "(sub)"
        np = bool(sub.get("is_non_prelimed"))
        lwn.append((ln, np))
        (nonprelim if np else prelim).append(ln)
    flags = rs.derive_tracker_flags(lwn)

    unb = (sb.table("release_unbilled_entries").select("*")
           .eq("release_tracker_id", str(tracker_id)).order("sort_order").execute()).data or []

    wb = Workbook()
    ws = wb.active
    ws.title = t.get("period") or "Tracker"
    ws.column_dimensions["A"].width = 34
    for c in "BCDE":
        ws.column_dimensions[c].width = 16

    def label(row, text, col="A", bold=True):
        cell = ws.cell(row=row, column=ord(col) - 64, value=text)
        if bold:
            cell.font = _BOLD
        return cell

    def money(row, col, value_or_formula):
        cell = ws.cell(row=row, column=ord(col) - 64, value=value_or_formula)
        cell.number_format = MONEY
        return cell

    # ── Header block ──
    label(1, "PROJECT:"); ws["C1"] = project.get("name", "")
    label(2, "PROJECT NO:"); ws["C2"] = project.get("project_no", "")
    label(3, "APPLICATION NO:"); ws["C3"] = t.get("period", "")
    label(4, "PERIOD TO:"); ws["C4"] = _period_to(t.get("period", ""))
    label(1, "Requested releases:", col="D"); ws["F1"] = "Yes" if flags["requested_releases"] else "No"
    label(2, "Verified releases:", col="D"); ws["F2"] = "Yes" if flags["verified_releases"] else "No"
    label(3, "Approved:", col="D"); ws["F3"] = "Yes" if t.get("approved") else "No"
    label(4, "Sent to GC:", col="D"); ws["F4"] = "Yes" if flags["sent_to_gc"] else "No"

    label(6, "Invoice Amount"); money(6, "D", float(t["invoice_amount"]) if t.get("invoice_amount") else 0)
    label(7, "Conditional Through Date"); ws["D7"] = str(t.get("conditional_through_date") or "")

    # ── Subs / Vendors table ──
    r = 9
    hdr = ["Subs / Vendors", "Billed", "Check", "Difference"]
    for i, h in enumerate(hdr):
        c = ws.cell(row=r, column=1 + i, value=h)
        c.font = _BOLD
        c.fill = _SHADE
    r += 1
    subs_start = r
    for ln in prelim:
        ws.cell(row=r, column=1, value=ln["sub_name"])
        money(r, "B", float(ln.get("billed_amount") or 0))
        money(r, "C", float(ln.get("check_amount") or 0))
        money(r, "D", f"=B{r}-C{r}")
        r += 1
    subs_end = r - 1
    subs_total_row = r
    label(r, "Subs / Vendors Total")
    if prelim:
        money(r, "B", f"=SUM(B{subs_start}:B{subs_end})")
        money(r, "C", f"=SUM(C{subs_start}:C{subs_end})")
        money(r, "D", f"=SUM(D{subs_start}:D{subs_end})")
    else:
        for col in "BCD":
            money(r, col, 0)
    r += 2

    ferro_total_row = r
    label(r, "Ferrocrete Builders, Inc. Total")
    money(r, "C", f"=D6-C{subs_total_row}")   # Invoice − Subs/Vendors checks
    r += 2

    # ── Non-Prelimed Bills ──
    label(r, "Non-Prelimed Bills"); ws.cell(row=r, column=1).fill = _SHADE
    r += 1
    np_start = r
    if nonprelim:
        for ln in nonprelim:
            ws.cell(row=r, column=1, value=ln["sub_name"])
            money(r, "C", float(ln.get("check_amount") or 0))
            r += 1
    else:
        r += 1
    np_end = r - 1
    np_total_row = r
    label(r, "Non-Prelimed Total")
    money(r, "C", f"=SUM(C{np_start}:C{np_end})")
    r += 2

    # ── Previous Month(s) Unbilled ──
    label(r, "Previous Month(s) Unbilled Balance Due"); ws.cell(row=r, column=1).fill = _SHADE
    r += 1
    unb_start = r
    for u in (unb or [{"amount": 0}]):
        ws.cell(row=r, column=1, value=(u.get("description") or ""))
        money(r, "C", float(u.get("amount") or 0))
        r += 1
    unb_end = r - 1
    unb_total_row = r
    label(r, "Previous Month(s) Unbilled Total")
    money(r, "C", f"=SUM(C{unb_start}:C{unb_end})")
    r += 2

    # ── Ferrocrete Net ──
    net_row = r
    label(r, "Ferrocrete Builders, Inc. Net")
    money(r, "C", f"=C{ferro_total_row}-C{np_total_row}-C{unb_total_row}")
    r += 2

    # ── Buildertrend reconciliation ──
    label(r, "Buildertrend Bills Total")
    money(r, "D", float(t.get("buildertrend_total") or 0))
    bt_row = r
    r += 1
    label(r, "Previous Month(s) Balance Due")
    money(r, "D", f"=C{unb_total_row}")
    prevbal_row = r
    r += 1
    label(r, "Less Misc. Field Expenses")
    money(r, "D", float(t.get("less_misc_field_expenses") or 0))
    misc_row = r
    money(r, "C", f"=D{bt_row}+D{prevbal_row}-D{misc_row}")   # Buildertrend side
    r += 2

    label(r, "Spreadsheet Bills Total")
    money(r, "C", f"=C{subs_total_row}+C{np_total_row}+C{unb_total_row}")
    ss_row = r
    r += 1
    label(r, "Discrepancy")
    money(r, "C", f"=C{misc_row}-C{ss_row}")
    r += 1

    for row in range(1, r):
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{project.get('project_no','')}_{project.get('name','')}_" \
            f"{t.get('period','')}_Release_Tracker.xlsx".replace("/", "-")
    return buf.getvalue(), fname
