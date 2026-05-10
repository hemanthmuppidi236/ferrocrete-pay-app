"""
Pay app Excel/PDF generation.

Bridges the database state (pay_apps + billings + sov_lines + change_orders) to the
existing `payapp_engine.py` from Demo 1, which operates on Excel files.

Strategy:
  1. Load pay app + billings + project + SOV + COs from DB
  2. Build an in-memory "project" data structure matching what payapp_engine expects
  3. Use a copy of the PayApp_Template.xlsx, populate via openpyxl directly
     (we don't reuse fabricate_projects.py because it does too much)
  4. Save the resulting bytes
  5. Upload to Supabase Storage
  6. Return a signed URL
"""

from io import BytesIO
from pathlib import Path
from datetime import datetime, timezone, date
from decimal import Decimal
from typing import List
import calendar
import shutil
import tempfile

from openpyxl import load_workbook

from .config import settings
from .supabase_client import get_service_client
from .pay_app_math import calculate_pay_app_totals
from . import storage as storage_helpers


# Path to the Excel template inside the backend repo
TEMPLATE_DIR = Path(__file__).resolve().parents[2] / "engines" / "templates"
PAY_APP_TEMPLATE = TEMPLATE_DIR / "PayApp_Template.xlsx"


# ─── Cell map (matches Demo 1's payapp_engine) ────────────────────────
CELL_PROJECT_NAME = "C1"
CELL_PROJECT_NO = "I3"
CELL_APP_NO = "I2"
CELL_APP_DATE = "I4"
CELL_PERIOD_TO = "I5"

SHEET_702 = "702"
SHEET_G703 = "G703"

SOV_BODY_START = 15
SOV_BODY_END = 72
CO_BODY_START = 74
CO_BODY_END = 76
GRAND_TOTAL_ROW = 78

S702_RETENTION_RATE = "C27"
S702_PREVIOUS_CERTS = "G29"


def generate_pay_app_excel(pay_app_id: str) -> bytes:
    """Generate the pay app Excel file. Returns bytes."""
    if not PAY_APP_TEMPLATE.exists():
        raise FileNotFoundError(
            f"Pay app template not found at {PAY_APP_TEMPLATE}. "
            "Make sure backend/engines/templates/PayApp_Template.xlsx exists."
        )

    sb = get_service_client()

    # Load pay app + project
    pa_res = sb.table("pay_apps").select("*, projects(*)").eq("id", str(pay_app_id)).limit(1).execute()
    if not pa_res.data:
        raise ValueError(f"Pay app {pay_app_id} not found")
    pa = pa_res.data[0]
    project = pa["projects"]

    sov_res = sb.table("sov_lines").select("*").eq("project_id", project["id"]).order("sort_order").execute()
    co_res = sb.table("change_orders").select("*").eq("project_id", project["id"]).order("co_no").execute()
    bil_res = sb.table("pay_app_billings").select("*").eq("pay_app_id", str(pay_app_id)).execute()

    # Index billings by sov_line_id and change_order_id for quick lookup
    sov_billings = {b["sov_line_id"]: b for b in bil_res.data if b.get("sov_line_id")}
    co_billings = {b["change_order_id"]: b for b in bil_res.data if b.get("change_order_id")}

    # Load template into memory
    with open(PAY_APP_TEMPLATE, "rb") as f:
        template_bytes = f.read()
    wb = load_workbook(BytesIO(template_bytes))
    g703 = wb[SHEET_G703]
    s702 = wb[SHEET_702]

    # Header
    g703[CELL_PROJECT_NAME] = project["name"]
    g703[CELL_PROJECT_NO] = project["project_no"]
    g703[CELL_APP_NO] = pa["app_no"]
    g703[CELL_APP_DATE] = datetime.now(timezone.utc).date()
    if pa.get("period_to"):
        period_to = pa["period_to"]
        if isinstance(period_to, str):
            period_to = date.fromisoformat(period_to)
        g703[CELL_PERIOD_TO] = period_to

    # Retention rate on 702!C27
    s702[S702_RETENTION_RATE] = float(project["retention_rate"])

    # Previous certificates (carried from prior pay apps)
    s702[S702_PREVIOUS_CERTS] = float(pa.get("previous_certificates") or 0)

    # SOV body — rows 15 to 72 (max 58 lines)
    if len(sov_res.data) > (SOV_BODY_END - SOV_BODY_START + 1):
        raise ValueError(f"Too many SOV lines ({len(sov_res.data)}); template supports {SOV_BODY_END - SOV_BODY_START + 1}")
    for i, sov in enumerate(sov_res.data):
        r = SOV_BODY_START + i
        g703.cell(row=r, column=1, value=sov.get("item_no") or str(i + 1))   # A: item #
        g703.cell(row=r, column=2, value=sov.get("description"))             # B: description
        g703.cell(row=r, column=3, value=float(sov.get("scheduled_value") or 0))  # C: scheduled value

        b = sov_billings.get(sov["id"])
        if b:
            g703.cell(row=r, column=4, value=float(b.get("previous_work") or 0))  # D
            g703.cell(row=r, column=5, value=float(b.get("this_period_work") or 0))  # E
            g703.cell(row=r, column=6, value=float(b.get("materials_stored") or 0))  # F

    # CO body — rows 74 to 76 (max 3 COs)
    approved_cos = [co for co in co_res.data if co["status"] == "approved"]
    if len(approved_cos) > (CO_BODY_END - CO_BODY_START + 1):
        # Lump remainder; for now warn
        pass
    for i, co in enumerate(approved_cos[: CO_BODY_END - CO_BODY_START + 1]):
        r = CO_BODY_START + i
        g703.cell(row=r, column=1, value=co.get("co_no"))
        g703.cell(row=r, column=2, value=co.get("description"))
        g703.cell(row=r, column=3, value=float(co.get("amount") or 0))
        b = co_billings.get(co["id"])
        if b:
            g703.cell(row=r, column=4, value=float(b.get("previous_work") or 0))
            g703.cell(row=r, column=5, value=float(b.get("this_period_work") or 0))
            g703.cell(row=r, column=6, value=float(b.get("materials_stored") or 0))

    # Save to bytes
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_and_store_pay_app_excel(pay_app_id: str) -> dict:
    """Generate the Excel, upload to storage, mark pay_app row, return metadata."""
    sb = get_service_client()
    pa_res = sb.table("pay_apps").select("*, projects(project_no, name)").eq("id", str(pay_app_id)).limit(1).execute()
    if not pa_res.data:
        raise ValueError(f"Pay app {pay_app_id} not found")
    pa = pa_res.data[0]
    project = pa["projects"]

    excel_bytes = generate_pay_app_excel(pay_app_id)
    path = storage_helpers.make_pay_app_excel_path(
        pa["period"], project["project_no"], project["name"],
    )
    storage_helpers.upload_bytes(
        settings.bucket_pay_apps,
        path,
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    now = datetime.now(timezone.utc).isoformat()
    sb.table("pay_apps").update({
        "excel_file_path": path,
        "excel_generated_at": now,
    }).eq("id", str(pay_app_id)).execute()

    return {
        "file_path": path,
        "download_url": storage_helpers.signed_url(settings.bucket_pay_apps, path),
        "generated_at": now,
    }
