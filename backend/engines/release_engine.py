"""
Release Engine — manages monthly release tracker files and waiver folders.

Each draw period creates:
  03 Releases/{period}/{project_folder}/
    {project_name}_Release_Tracker.xlsx   ← carries history forward + new sheet
    CP/   ← empty folder for received conditional progress waivers
    UP/   ← empty folder for received unconditional progress waivers
    CF/   ← empty folder for received conditional final waivers
    UF/   ← empty folder for received unconditional final waivers

The tracker file accumulates one sheet per draw period (matching the format
the user already uses), plus a Data sheet with dropdown sources.

Column layout (matches the user's existing tracker exactly):
  A: Sub/Vendor name (or "Sub-Tier" marker)
  B: Sub-Tier sub-name (when A == "Sub-Tier")
  C: (merged with B for sub-tier rows)
  D: Billed Amount
  E: Check Amount
  F: Release Type (CP/CF/UP/UF)
  G: Release Exception (Y/N/N/A)
  H: Previous Month UP/UF Received status
  J: Computed difference (D - E for parents; D - E for some children)

The sub-block size is variable. The engine maintains:
  - Sub block: rows SUB_START through SUB_END (where SUB_END floats with sub count)
  - "Subs/Vendors Total" row: SUB_END + 2
  - All downstream rows (Ferrocrete Total, Non-Prelimed, etc.) shift accordingly
"""

import re
import shutil
from copy import copy
from pathlib import Path
from datetime import date
import calendar

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.copier import WorksheetCopy

# ─── Cell map (matches user's existing tracker) ─────────────────────
HDR_PROJECT_NAME = 'C1'
HDR_PROJECT_NO = 'C2'
HDR_APP_NO = 'C3'
HDR_PERIOD_TO = 'C4'
HDR_INVOICE_AMOUNT = 'D11'
HDR_CONDITIONAL_THROUGH = 'D14'

SUB_BLOCK_START = 16  # First sub row
# The block extends as needed. After last sub, there's 1 empty row,
# then "Subs/Vendors Total" at SUB_END + 2.

# Anchored sections (these shift if sub block grows):
# After "Subs/Vendors Total" → 1 empty row → "Ferrocrete Total"
# Then 1 empty row → "Non-Prelimed Bills" header, 3 vendor rows, "Non-Prelimed Total"
# Then 1 empty row → "Previous Month(s) Unbilled Balance Due" header, 5 vendor rows, total
# Then 1 empty row → "Ferrocrete Net"
# Then 1 empty row → "Buildertrend Bills Total" / "Previous Month Balance Due" / "Less Misc"
# Then 1 empty row → "Spreadsheet Bills Total" / "Discrepancy"

NON_PRELIM_VENDORS = 3   # configurable; user's file has 3
PREV_UNBILLED_VENDORS = 5  # user's file has 5


def period_to_dates(period_str):
    """'26-04' -> (date(2026,4,1), date(2026,4,30))"""
    yy, mm = period_str.split('-')
    year = 2000 + int(yy)
    month = int(mm)
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def prev_period(period_str):
    """'26-04' -> '26-03'"""
    yy, mm = period_str.split('-')
    yy, mm = int(yy), int(mm)
    mm -= 1
    if mm == 0:
        mm = 12
        yy -= 1
    return f"{yy:02d}-{mm:02d}"


def safe_set(ws, row, col, value):
    """Skip merged cells gracefully."""
    cell = ws.cell(row=row, column=col)
    # If cell is part of a merged range and not the anchor, skip
    for mr in ws.merged_cells.ranges:
        if (mr.min_row <= row <= mr.max_row and
            mr.min_col <= col <= mr.max_col and
            (row != mr.min_row or col != mr.min_col)):
            return
    cell.value = value


def _copy_cell_style(src_cell, dst_cell):
    """Copy formatting from one cell to another."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def _unmerge_range(ws, top, bottom):
    """Remove all merged cell ranges that fall within [top, bottom] rows."""
    to_remove = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= top and mr.max_row <= bottom:
            to_remove.append(str(mr))
    for r in to_remove:
        ws.unmerge_cells(r)


def _clear_range(ws, top, bottom, cols=range(1, 12)):
    """Clear cell values AND formatting in rows [top, bottom] for the given columns.
    Removes merges first.
    """
    from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
    _unmerge_range(ws, top, bottom)
    default_font = Font()
    default_fill = PatternFill()
    default_border = Border()
    default_alignment = Alignment()
    default_protection = Protection()
    for r in range(top, bottom + 1):
        for c in cols:
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.font = default_font
            cell.fill = default_fill
            cell.border = default_border
            cell.alignment = default_alignment
            cell.number_format = 'General'
            cell.protection = default_protection


def populate_sheet(ws, project_info, subs, period, invoice_amount=None):
    """
    Populate a release tracker sheet. Strategy:
      1. Write project header
      2. Find existing "Subs/Vendors Total" to bound the block
      3. Strip the entire sub block + downstream sections
      4. Rewrite sub block from scratch with correct merges per row type
      5. Rewrite downstream sections with formulas re-pointed to new row numbers
    """
    period_start, period_end = period_to_dates(period)

    # ─── Header ───
    ws[HDR_PROJECT_NAME] = project_info['name']
    ws[HDR_PROJECT_NO] = project_info['project_no']
    ws[HDR_APP_NO] = period
    ws[HDR_PERIOD_TO] = period_end
    # Always clear D11 first so any inherited value from the source-clone sheet
    # doesn't leak through when invoice_amount is None (e.g., empty draft pay app).
    ws[HDR_INVOICE_AMOUNT] = invoice_amount  # None or float
    ws[HDR_CONDITIONAL_THROUGH] = f"Conditional Through Date {period_end.strftime('%m-%d-%y')}"

    # ─── Separate prime and non-prelimed subs ───
    prime_subs = [s for s in subs if not s.get('non_prelimed')]
    non_prelim_subs = [s for s in subs if s.get('non_prelimed')]

    # ─── Build flat row plan for sub block ───
    sub_rows = []  # list of (kind, sub_or_subtier, parent_idx_or_None)
    for sub in prime_subs:
        parent_idx = len(sub_rows)
        sub_rows.append(('parent', sub, None))
        for sub_tier in sub.get('sub_tiers', []) or []:
            sub_rows.append(('subtier', sub_tier, parent_idx))

    # Find the entire range to nuke and rebuild — from SUB_BLOCK_START through Discrepancy
    discrepancy_row = None
    for r in range(SUB_BLOCK_START, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == 'Discrepancy':
            discrepancy_row = r
            break

    # Capture template row styles BEFORE we clear anything
    # Use row 16 as the parent-sub style template, row 22 as sub-tier template if present
    parent_styles = _capture_row_styles(ws, SUB_BLOCK_START)
    subtier_styles = None
    # Try to find a sub-tier row in the template to copy styles from (B22:D22 merged in template)
    for mr in ws.merged_cells.ranges:
        if mr.min_col == 2 and mr.max_col == 4 and mr.min_row == mr.max_row:
            # B?:D? merge — that's a sub-tier row pattern
            subtier_styles = _capture_row_styles(ws, mr.min_row)
            break

    # Capture section-row styles by searching for known labels in column A.
    # These are the rows whose formatting (borders, fills, fonts) we want to preserve.
    section_styles = {}
    section_labels = [
        'Subs/Vendors Total',
        'Ferrocrete Builders, Inc. Total',
        'Non-Prelimed Bills',
        'Non-Prelimed Total',
        'Previous Month(s) Unbilled Balance Due',
        'Previous Month(s) Unbilled Balance Due Total',
        'Ferrocrete Builders, Inc. Net',
        'Buildertrend Bills Total',
        'Previous Month(s) Balance Due',
        'Less Misc. Field Expenses',
        'Spreadsheet Bills Total',
        'Discrepancy',
    ]
    # Also capture vendor-row style (a non-prelim or prev-unbilled row in original)
    vendor_row_style = None
    for r in range(SUB_BLOCK_START, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if a_val in section_labels:
            section_styles[a_val] = _capture_row_styles(ws, r)
        # Empty rows just below a section header are typically vendor rows
        # (e.g., 3 vendors under Non-Prelimed Bills). Use the first one as vendor template.
        if vendor_row_style is None and a_val is None:
            # Check if any of the row's cells have borders (i.e., it's a vendor row, not pure spacer)
            for c in range(1, 12):
                cell = ws.cell(row=r, column=c)
                if cell.border and any(getattr(cell.border, s).style for s in ['top', 'bottom', 'left', 'right'] if getattr(cell.border, s)):
                    vendor_row_style = _capture_row_styles(ws, r)
                    break

    # Compute new layout rows
    n_sub_rows = len(sub_rows) if sub_rows else 1  # at least 1 placeholder row so block isn't empty
    sub_block_end = SUB_BLOCK_START + n_sub_rows - 1
    spacer1 = sub_block_end + 1
    sub_total_row = sub_block_end + 2
    spacer2 = sub_total_row + 1
    ferro_total_row = sub_total_row + 2
    spacer3 = ferro_total_row + 1
    non_prelim_header_row = ferro_total_row + 2
    non_prelim_first = non_prelim_header_row + 1
    non_prelim_last = non_prelim_header_row + NON_PRELIM_VENDORS
    non_prelim_total_row = non_prelim_last + 1
    spacer4 = non_prelim_total_row + 1
    prev_unbilled_header_row = non_prelim_total_row + 2
    prev_unbilled_first = prev_unbilled_header_row + 1
    prev_unbilled_last = prev_unbilled_header_row + PREV_UNBILLED_VENDORS
    prev_unbilled_total_row = prev_unbilled_last + 1
    spacer5 = prev_unbilled_total_row + 1
    ferro_net_row = prev_unbilled_total_row + 2
    spacer6 = ferro_net_row + 1
    buildertrend_row = ferro_net_row + 2
    prev_balance_row = buildertrend_row + 1
    less_misc_row = buildertrend_row + 2
    spacer7 = less_misc_row + 1
    spreadsheet_total_row = less_misc_row + 2
    new_discrepancy_row = spreadsheet_total_row + 1

    # Clear the whole working area (from SUB_BLOCK_START down to whichever is bigger:
    # the existing discrepancy row or our new discrepancy row)
    nuke_to = max(discrepancy_row or 0, new_discrepancy_row, ws.max_row)
    _clear_range(ws, SUB_BLOCK_START, nuke_to)

    # ─── Write sub block ───
    if sub_rows:
        for i, (kind, sub, parent_idx) in enumerate(sub_rows):
            r = SUB_BLOCK_START + i
            if kind == 'parent':
                # Merge A:C
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
                ws.cell(row=r, column=1).value = sub['name']
                ws.cell(row=r, column=4).value = sub.get('billed', 0)
                # E formula or value depending on sub-tier presence
                sub_tier_indices = [j for j, (k, s, p) in enumerate(sub_rows) if p == i]
                if sub_tier_indices:
                    sub_tier_row_nums = [SUB_BLOCK_START + j for j in sub_tier_indices]
                    formula = f"=D{r}-" + "-".join(f"E{tr}" for tr in sub_tier_row_nums)
                    ws.cell(row=r, column=5).value = formula
                else:
                    ws.cell(row=r, column=5).value = sub.get('check', 0)
                ws.cell(row=r, column=6).value = sub.get('release_type')
                ws.cell(row=r, column=7).value = sub.get('exception')
                ws.cell(row=r, column=8).value = sub.get('prev_month_status')
                ws.cell(row=r, column=10).value = f'=IF(D{r}="",E{r},D{r}-E{r})'
                _apply_row_styles(ws, r, parent_styles)
            else:  # subtier
                # Merge B:D for the sub-tier name
                ws.cell(row=r, column=1).value = 'Sub-Tier'
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
                ws.cell(row=r, column=2).value = sub['name']
                # Sub-tier has no D (billed) — only E (check)
                ws.cell(row=r, column=5).value = sub.get('check', 0)
                ws.cell(row=r, column=6).value = sub.get('release_type')
                ws.cell(row=r, column=7).value = sub.get('exception')
                ws.cell(row=r, column=8).value = sub.get('prev_month_status')
                ws.cell(row=r, column=10).value = f'=IF(D{r}="",E{r},D{r}-E{r})'
                _apply_row_styles(ws, r, subtier_styles or parent_styles)
    else:
        # Empty placeholder row (so the sheet doesn't look broken)
        r = SUB_BLOCK_START
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        _apply_row_styles(ws, r, parent_styles)

    # ─── Subs/Vendors Total ───
    last_sub_data_row = SUB_BLOCK_START + len(sub_rows) - 1 if sub_rows else SUB_BLOCK_START
    ws.merge_cells(start_row=sub_total_row, start_column=1, end_row=sub_total_row, end_column=4)
    ws.cell(row=sub_total_row, column=1).value = 'Subs/Vendors Total'
    if sub_rows:
        ws.cell(row=sub_total_row, column=5).value = f'=SUM(E{SUB_BLOCK_START}:E{last_sub_data_row})'
    else:
        ws.cell(row=sub_total_row, column=5).value = 0
    _apply_row_styles(ws, sub_total_row, section_styles.get('Subs/Vendors Total'))

    # ─── Ferrocrete Builders, Inc. Total ───
    ws.merge_cells(start_row=ferro_total_row, start_column=1, end_row=ferro_total_row, end_column=4)
    ws.cell(row=ferro_total_row, column=1).value = 'Ferrocrete Builders, Inc. Total'
    ws.cell(row=ferro_total_row, column=5).value = f'=D11-E{sub_total_row}'
    _apply_row_styles(ws, ferro_total_row, section_styles.get('Ferrocrete Builders, Inc. Total'))

    # ─── Non-Prelimed Bills section ───
    ws.merge_cells(start_row=non_prelim_header_row, start_column=1, end_row=non_prelim_header_row, end_column=4)
    ws.cell(row=non_prelim_header_row, column=1).value = 'Non-Prelimed Bills'
    _apply_row_styles(ws, non_prelim_header_row, section_styles.get('Non-Prelimed Bills'))
    for i in range(NON_PRELIM_VENDORS):
        r = non_prelim_first + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        if i < len(non_prelim_subs):
            ws.cell(row=r, column=1).value = non_prelim_subs[i]['name']
            ws.cell(row=r, column=5).value = non_prelim_subs[i].get('check', 0)
        else:
            ws.cell(row=r, column=5).value = 0
        _apply_row_styles(ws, r, vendor_row_style)
    ws.merge_cells(start_row=non_prelim_total_row, start_column=1, end_row=non_prelim_total_row, end_column=4)
    ws.cell(row=non_prelim_total_row, column=1).value = 'Non-Prelimed Total'
    ws.cell(row=non_prelim_total_row, column=5).value = f'=SUM(E{non_prelim_first}:E{non_prelim_last})'
    _apply_row_styles(ws, non_prelim_total_row, section_styles.get('Non-Prelimed Total'))

    # ─── Previous Month(s) Unbilled Balance Due ───
    ws.merge_cells(start_row=prev_unbilled_header_row, start_column=1, end_row=prev_unbilled_header_row, end_column=4)
    ws.cell(row=prev_unbilled_header_row, column=1).value = 'Previous Month(s) Unbilled Balance Due'
    _apply_row_styles(ws, prev_unbilled_header_row, section_styles.get('Previous Month(s) Unbilled Balance Due'))
    for i in range(PREV_UNBILLED_VENDORS):
        r = prev_unbilled_first + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(row=r, column=5).value = 0
        _apply_row_styles(ws, r, vendor_row_style)
    ws.merge_cells(start_row=prev_unbilled_total_row, start_column=1, end_row=prev_unbilled_total_row, end_column=4)
    ws.cell(row=prev_unbilled_total_row, column=1).value = 'Previous Month(s) Unbilled Balance Due Total'
    ws.cell(row=prev_unbilled_total_row, column=5).value = f'=SUM(E{prev_unbilled_first}:E{prev_unbilled_last})'
    _apply_row_styles(ws, prev_unbilled_total_row, section_styles.get('Previous Month(s) Unbilled Balance Due Total'))

    # ─── Ferrocrete Builders, Inc. Net ───
    ws.merge_cells(start_row=ferro_net_row, start_column=1, end_row=ferro_net_row, end_column=4)
    ws.cell(row=ferro_net_row, column=1).value = 'Ferrocrete Builders, Inc. Net'
    ws.cell(row=ferro_net_row, column=5).value = f'=E{ferro_total_row}-E{non_prelim_total_row}-E{prev_unbilled_total_row}'
    _apply_row_styles(ws, ferro_net_row, section_styles.get('Ferrocrete Builders, Inc. Net'))

    # ─── Buildertrend reconciliation ───
    ws.merge_cells(start_row=buildertrend_row, start_column=1, end_row=buildertrend_row, end_column=3)
    ws.cell(row=buildertrend_row, column=1).value = 'Buildertrend Bills Total'
    _apply_row_styles(ws, buildertrend_row, section_styles.get('Buildertrend Bills Total'))
    ws.merge_cells(start_row=prev_balance_row, start_column=1, end_row=prev_balance_row, end_column=3)
    ws.cell(row=prev_balance_row, column=1).value = 'Previous Month(s) Balance Due'
    ws.cell(row=prev_balance_row, column=4).value = f'=E{prev_unbilled_total_row}'
    _apply_row_styles(ws, prev_balance_row, section_styles.get('Previous Month(s) Balance Due'))
    ws.merge_cells(start_row=less_misc_row, start_column=1, end_row=less_misc_row, end_column=3)
    ws.cell(row=less_misc_row, column=1).value = 'Less Misc. Field Expenses'
    ws.cell(row=less_misc_row, column=4).value = 0
    ws.cell(row=less_misc_row, column=5).value = f'=D{buildertrend_row}+D{prev_balance_row}-D{less_misc_row}'
    _apply_row_styles(ws, less_misc_row, section_styles.get('Less Misc. Field Expenses'))

    # ─── Spreadsheet Bills Total / Discrepancy ───
    ws.merge_cells(start_row=spreadsheet_total_row, start_column=1, end_row=spreadsheet_total_row, end_column=4)
    ws.cell(row=spreadsheet_total_row, column=1).value = 'Spreadsheet Bills Total'
    ws.cell(row=spreadsheet_total_row, column=5).value = f'=E{sub_total_row}+E{non_prelim_total_row}+E{prev_unbilled_total_row}'
    _apply_row_styles(ws, spreadsheet_total_row, section_styles.get('Spreadsheet Bills Total'))
    ws.merge_cells(start_row=new_discrepancy_row, start_column=1, end_row=new_discrepancy_row, end_column=4)
    ws.cell(row=new_discrepancy_row, column=1).value = 'Discrepancy'
    ws.cell(row=new_discrepancy_row, column=5).value = f'=E{less_misc_row}-E{spreadsheet_total_row}'
    _apply_row_styles(ws, new_discrepancy_row, section_styles.get('Discrepancy'))

    # ─── Trim any leftover rows past Discrepancy ───
    if ws.max_row > new_discrepancy_row:
        rows_to_delete = ws.max_row - new_discrepancy_row
        ws.delete_rows(new_discrepancy_row + 1, amount=rows_to_delete)


def _capture_row_styles(ws, row):
    """Snapshot per-column style for a row."""
    styles = {}
    for col in range(1, 12):
        cell = ws.cell(row=row, column=col)
        if cell.has_style:
            styles[col] = {
                'font': copy(cell.font),
                'fill': copy(cell.fill),
                'border': copy(cell.border),
                'alignment': copy(cell.alignment),
                'number_format': cell.number_format,
                'protection': copy(cell.protection),
            }
    return styles


def _apply_row_styles(ws, row, styles):
    """Apply captured styles to a row."""
    if not styles:
        return
    for col, s in styles.items():
        cell = ws.cell(row=row, column=col)
        cell.font = copy(s['font'])
        cell.fill = copy(s['fill'])
        cell.border = copy(s['border'])
        cell.alignment = copy(s['alignment'])
        cell.number_format = s['number_format']
        cell.protection = copy(s['protection'])


def copy_sheet_to_workbook(src_ws, dst_wb, new_name):
    """Copy a worksheet from one workbook to another with formatting."""
    dst_ws = dst_wb.create_sheet(title=new_name)
    copier = WorksheetCopy(source_worksheet=src_ws, target_worksheet=dst_ws)
    copier.copy_worksheet()
    return dst_ws


def add_period_sheet(filepath, period, project_info, subs, invoice_amount=None,
                     template_path=None, prior_period_for_status=None):
    """
    Add a new sheet for the given period to the release tracker workbook.

    If the file doesn't exist, create it from the template.
    If it exists, copy the TEMPLATE sheet structure into a new sheet named `period`.

    The new sheet is placed in chronological order (period names sort lexically: 25-06 < 26-04).
    """
    filepath = Path(filepath)

    if not filepath.exists():
        # Create from template
        if template_path is None:
            raise ValueError("template_path required when filepath doesn't exist")
        filepath.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(template_path, filepath)

    wb = load_workbook(filepath)

    # If a sheet for this period already exists, leave it alone
    if period in wb.sheetnames:
        wb.close()
        return False

    # Find the TEMPLATE sheet (or use the most recent period sheet) to copy from
    src_sheet_name = 'TEMPLATE'
    if src_sheet_name not in wb.sheetnames:
        # Use the most recent period sheet
        period_sheets = [s for s in wb.sheetnames if re.match(r'^\d{2}-\d{2}$', s)]
        if not period_sheets:
            wb.close()
            raise ValueError("No TEMPLATE sheet and no period sheets found")
        src_sheet_name = sorted(period_sheets)[-1]

    src_ws = wb[src_sheet_name]

    # Copy via WorksheetCopy
    new_ws = wb.copy_worksheet(src_ws)
    new_ws.title = period

    # If source was TEMPLATE and there are no other period sheets yet, delete TEMPLATE
    # (we keep it for the very first creation, then remove)
    template_sheet = wb['TEMPLATE'] if 'TEMPLATE' in wb.sheetnames else None
    period_sheets = [s for s in wb.sheetnames if re.match(r'^\d{2}-\d{2}$', s)]

    # Order sheets: all period sheets in chronological order, Data last
    sorted_periods = sorted(period_sheets)
    sheet_order = sorted_periods + ['Data'] if 'Data' in wb.sheetnames else sorted_periods

    # Reorder via _sheets list manipulation (preserves objects)
    new_order = []
    name_to_sheet = {ws.title: ws for ws in wb.worksheets}
    for name in sheet_order:
        if name in name_to_sheet:
            new_order.append(name_to_sheet[name])
    # Tack on any sheets we missed (e.g., TEMPLATE if still present)
    for ws in wb.worksheets:
        if ws not in new_order:
            new_order.append(ws)
    wb._sheets = new_order

    # Populate the new sheet
    populate_sheet(new_ws, project_info, subs, period, invoice_amount)

    # Apply unlocked-cell protection (allow editing sub list, billed amounts, check amounts)
    apply_protection(new_ws)

    # Remove TEMPLATE sheet now that we have at least one real period sheet
    if 'TEMPLATE' in wb.sheetnames:
        del wb['TEMPLATE']

    wb.save(filepath)
    wb.close()
    return True


def apply_protection(ws):
    """Set cell-level protection: unlock columns A, B, D, E, F, G, H in the sub block.
    Keep formulas (J column, totals) locked. No password.
    """
    # Find the "Discrepancy" row to determine the editable area
    discrepancy_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == 'Discrepancy':
            discrepancy_row = r
            break

    end_row = discrepancy_row if discrepancy_row else ws.max_row

    # Unlock user-editable cells: in sub block rows, columns A, B, D, E, F, G, H
    # Header (rows 1-15) editable: C1-C4, D11
    UNLOCK_HEADER = ['C1', 'C2', 'C3', 'C4', 'D11']
    for coord in UNLOCK_HEADER:
        cell = ws[coord]
        new_prot = copy(cell.protection)
        new_prot.locked = False
        cell.protection = new_prot

    # Sub block: rows SUB_BLOCK_START through "Discrepancy" — unlock A, B, D, E, F, G, H
    UNLOCK_COLS = [1, 2, 4, 5, 6, 7, 8]
    for r in range(SUB_BLOCK_START, end_row + 1):
        # Skip header/total rows where we want to keep things locked
        a_val = ws.cell(row=r, column=1).value
        if a_val and any(label in str(a_val) for label in [
            'Subs/Vendors Total', 'Ferrocrete Builders, Inc. Total',
            'Non-Prelimed Bills', 'Non-Prelimed Total',
            'Previous Month(s) Unbilled', 'Ferrocrete Builders, Inc. Net',
            'Buildertrend', 'Previous Month(s) Balance Due',
            'Less Misc', 'Spreadsheet Bills Total', 'Discrepancy'
        ]):
            continue
        for c in UNLOCK_COLS:
            cell = ws.cell(row=r, column=c)
            new_prot = copy(cell.protection)
            new_prot.locked = False
            cell.protection = new_prot

    # Apply sheet protection (no password)
    ws.protection = SheetProtection(
        sheet=True,
        password=None,
        formatCells=False, formatColumns=False, formatRows=False,
        insertColumns=False, insertRows=False,
        deleteColumns=False, deleteRows=False,
        selectLockedCells=False, selectUnlockedCells=False,
        sort=False, autoFilter=False, pivotTables=False,
    )


def get_subs_from_sheet(ws):
    """Read the sub list from an existing tracker sheet.

    Returns a list of dicts compatible with populate_sheet.
    """
    subs = []
    current_parent = None

    for r in range(SUB_BLOCK_START, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if a_val == 'Subs/Vendors Total':
            break
        if not a_val:
            continue  # empty row

        if a_val == 'Sub-Tier':
            sub_name = ws.cell(row=r, column=2).value
            if sub_name and current_parent:
                sub_tier = {
                    'name': sub_name.strip() if isinstance(sub_name, str) else sub_name,
                    'release_type': ws.cell(row=r, column=6).value,
                    'exception': ws.cell(row=r, column=7).value,
                    'billed': 0,
                    'check': 0,
                    'prev_month_status': ws.cell(row=r, column=8).value,
                }
                current_parent.setdefault('sub_tiers', []).append(sub_tier)
        else:
            sub = {
                'name': a_val.strip() if isinstance(a_val, str) else a_val,
                'release_type': ws.cell(row=r, column=6).value,
                'exception': ws.cell(row=r, column=7).value,
                'billed': 0,
                'check': 0,
                'prev_month_status': ws.cell(row=r, column=8).value,
                'sub_tiers': [],
            }
            subs.append(sub)
            current_parent = sub

    return subs


def carry_forward_subs(prior_filepath, prior_period):
    """Read sub list from the prior period's sheet, zero out amounts."""
    if not Path(prior_filepath).exists():
        return []
    wb = load_workbook(prior_filepath, data_only=False)
    if prior_period not in wb.sheetnames:
        # Use the most recent period
        period_sheets = sorted([s for s in wb.sheetnames if re.match(r'^\d{2}-\d{2}$', s)])
        if not period_sheets:
            wb.close()
            return []
        prior_period = period_sheets[-1]

    ws = wb[prior_period]
    subs = get_subs_from_sheet(ws)
    wb.close()

    # Zero out amounts but keep names, release types, exceptions
    for sub in subs:
        sub['billed'] = 0
        sub['check'] = 0
        sub['prev_month_status'] = None  # User decided not to auto-track for now
        for st in sub.get('sub_tiers', []) or []:
            st['billed'] = 0
            st['check'] = 0
            st['prev_month_status'] = None

    return subs


def pull_invoice_amount(payapp_filepath):
    """Read the 'Current Payment Due' from a pay app file.

    AIA G702/G703 chain:
      G703!G_row = D_row + E_row + F_row  (per line)
      G703!G78 = SUM(G15:G77) — grand total billed
      G703!J78 = SUM(J15:J76) — grand total retention
      702!G26 = G703!G78 — "Total Completed & Stored"
      702!G27 = G703!J78 — "Retainage"
      702!G28 = G26 - G27 — "Total Earned Less Retainage"
      702!G29 = (hardcoded by roll-forward) — "Less Previous Certificates"
      702!G30 = G28 - G29 — "Current Payment Due"

    Tries cached G30 first; if absent (file wasn't opened in Excel since last save),
    computes from the underlying D/E/F values on G703.
    """
    if not Path(payapp_filepath).exists():
        return None

    wb = load_workbook(payapp_filepath, data_only=True)
    if '702' not in wb.sheetnames or 'G703' not in wb.sheetnames:
        wb.close()
        return None

    s702 = wb['702']
    g703 = wb['G703']

    # Try cached G30 first
    cached = s702['G30'].value
    if cached is not None:
        try:
            wb.close()
            return float(cached)
        except (TypeError, ValueError):
            pass

    # Compute from underlying values
    try:
        # Sum D + E + F per row across G703 SOV (rows 15-72)
        sov_completed = 0.0
        for r in range(15, 73):
            d = g703.cell(row=r, column=4).value or 0
            e = g703.cell(row=r, column=5).value or 0
            f = g703.cell(row=r, column=6).value or 0
            try:
                sov_completed += float(d) + float(e) + float(f)
            except (TypeError, ValueError):
                continue

        # Sum D + E + F per row across G703 CO (rows 74-76; no retention)
        co_completed = 0.0
        for r in range(74, 77):
            d = g703.cell(row=r, column=4).value or 0
            e = g703.cell(row=r, column=5).value or 0
            f = g703.cell(row=r, column=6).value or 0
            try:
                co_completed += float(d) + float(e) + float(f)
            except (TypeError, ValueError):
                continue

        total_completed = sov_completed + co_completed

        # Retention rate from 702!C27
        rate = s702['C27'].value
        try:
            rate = float(rate) if rate else 0.10
        except (TypeError, ValueError):
            rate = 0.10
        total_retention = sov_completed * rate

        # Previous certificates (hardcoded by roll-forward)
        g29 = s702['G29'].value
        try:
            g29 = float(g29) if g29 else 0.0
        except (TypeError, ValueError):
            g29 = 0.0

        earned_less_ret = total_completed - total_retention
        current_pay_due = earned_less_ret - g29

        wb.close()
        # Allow None for empty drafts (avoid showing $0 when there's no real pull),
        # but allow positive values through.
        if current_pay_due < 1:  # essentially zero or negative
            return None
        return float(current_pay_due)
    except Exception:
        wb.close()
        return None


def scaffold_draw_folder(releases_root, period, project_folder_name):
    """Create the draw folder structure for a project.

    Creates:
        {releases_root}/{period}/{project_folder_name}/CP/
        {releases_root}/{period}/{project_folder_name}/UP/
        {releases_root}/{period}/{project_folder_name}/CF/
        {releases_root}/{period}/{project_folder_name}/UF/
    """
    base = Path(releases_root) / period / project_folder_name
    base.mkdir(parents=True, exist_ok=True)
    for waiver_type in ['CP', 'UP', 'CF', 'UF']:
        (base / waiver_type).mkdir(exist_ok=True)
    return base


def scaffold_release_for_project(
    releases_root, period, project_info, project_folder_name,
    template_path, prior_releases_root=None,
    payapp_filepath=None,
):
    """End-to-end: scaffold the draw folder + populate the tracker for one project.

    Args:
        releases_root: Path to '03 Releases' directory
        period: 'YY-MM' for the new draw
        project_info: dict with name, project_no
        project_folder_name: e.g., '25-05_712_Seagaze_Dr'
        template_path: path to ReleaseTracker_Template.xlsx
        prior_releases_root: same as releases_root (just for clarity); used to find prior
        payapp_filepath: path to the matching pay app file, for invoice auto-pull

    Returns the path to the new tracker file.
    """
    # 1. Create folder structure
    draw_folder = scaffold_draw_folder(releases_root, period, project_folder_name)

    # 2. Determine the new tracker file path
    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', project_info['name'])
    tracker_filename = f"{safe_name}_-_{period}_Release_Tracker.xlsx"
    tracker_filepath = draw_folder / tracker_filename

    # 3. If there's a prior tracker, copy it forward
    prior_period_str = prev_period(period)
    prior_tracker_path = None
    if prior_releases_root:
        prior_draw_folder = Path(prior_releases_root) / prior_period_str / project_folder_name
        if prior_draw_folder.exists():
            # Find the prior tracker file
            prior_files = list(prior_draw_folder.glob("*_Release_Tracker.xlsx"))
            if prior_files:
                prior_tracker_path = prior_files[0]
                shutil.copy(prior_tracker_path, tracker_filepath)

    # 4. Determine sub list — carry forward from prior, or empty
    if prior_tracker_path and prior_tracker_path.exists():
        subs = carry_forward_subs(prior_tracker_path, prior_period_str)
    else:
        subs = []

    # 5. Pull invoice amount from pay app if available
    invoice_amount = None
    if payapp_filepath:
        invoice_amount = pull_invoice_amount(payapp_filepath)

    # 6. Add the new period sheet
    add_period_sheet(
        filepath=tracker_filepath,
        period=period,
        project_info=project_info,
        subs=subs,
        invoice_amount=invoice_amount,
        template_path=template_path,
    )

    return tracker_filepath
