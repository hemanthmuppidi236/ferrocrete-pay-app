"""WI-6: tracker Excel export builds a valid workbook with formula totals."""

from io import BytesIO

from openpyxl import load_workbook

import app.core.tracker_export as te


class _Query:
    def __init__(self, rows):
        self._rows = list(rows)

    def select(self, *a, **k):
        return self

    def eq(self, f, v):
        self._rows = [r for r in self._rows if str(r.get(f)) == str(v)]
        return self

    def order(self, *a, **k):
        return self

    def limit(self, n):
        self._rows = self._rows[:n]
        return self

    def execute(self):
        return type("Res", (), {"data": self._rows})()


class _Client:
    def __init__(self, tables):
        self._tables = tables

    def table(self, name):
        return _Query(self._tables.get(name, []))


def _data():
    return {
        "release_trackers": [{
            "id": "t1", "project_id": "p1", "period": "26-06",
            "invoice_amount": "661319.60", "conditional_through_date": "2026-06-30",
            "buildertrend_total": "287647.87", "less_misc_field_expenses": "0",
            "approved": True,
        }],
        "projects": [{"id": "p1", "project_no": "25-20", "name": "A Street"}],
        "release_lines": [
            {"id": "l1", "release_tracker_id": "t1", "billed_amount": "967.43",
             "check_amount": "967.43", "bill_status": "received",
             "conditional_status": "sent_to_gc", "unconditional_status": "sent_to_gc",
             "subs": {"name": "Atlas Construction", "is_non_prelimed": False}},
            {"id": "l2", "release_tracker_id": "t1", "billed_amount": "0",
             "check_amount": "1610", "bill_status": "received",
             "conditional_status": "not_applicable", "unconditional_status": "not_applicable",
             "subs": {"name": "Non-Prelim Payments", "is_non_prelimed": True}},
        ],
        "release_unbilled_entries": [
            {"release_tracker_id": "t1", "description": "", "amount": "0", "sort_order": 0},
        ],
    }


def test_export_builds_and_has_formulas(monkeypatch):
    monkeypatch.setattr(te, "get_service_client", lambda: _Client(_data()))
    xlsx, fname = te.generate_tracker_export_xlsx("t1")
    assert fname.endswith("Release_Tracker.xlsx")
    assert "25-20" in fname

    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    # Header
    assert ws["C1"].value == "A Street"
    assert ws["C2"].value == "25-20"
    # Some cell must carry a SUM formula (totals are formulas, not baked values)
    formulas = [c.value for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    assert any("SUM(" in f for f in formulas)
    assert any(f.startswith("=B") and "-C" in f for f in formulas)  # Difference = B-C
    # Labels present
    labels = {c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)}
    assert "Subs / Vendors Total" in labels
    assert "Ferrocrete Builders, Inc. Net" in labels
    assert "Non-Prelimed Total" in labels
    assert "Discrepancy" in labels
